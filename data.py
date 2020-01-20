import os
from xlrd import open_workbook
from openpyxl import load_workbook


class get_data:
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.ready_data = []
        
       
        self.main_path = os.path.join(f'{self.file_path}') #// joins provided path and file stored in there(name provided by user)
        self.read = open_workbook(self.main_path)
        
        
        self.sheet = self.read.sheet_by_index(0)
        self.sheet.cell_value(0,0)

    def gather_data(self):
        for x in range(self.sheet.nrows):
            for y in range(self.sheet.ncols):
                cell = self.sheet.cell_value(x,y)
                self.ready_data.append(str(cell).split("/"))
        return self.ready_data

    def save_fails(self, id, reason, raport_name, raport_path):
        self.id = id
        self.reason = reason
        self.raport_name = raport_name
        self.raport_path = raport_path
    
        self.final_path = os.path.join(self.raport_path, f"{self.raport_name}.xlsx")
        found = False #// boolean for double check in excel

        if os.path.isfile(self.final_path) is False:
            self.modify = load_workbook(self.main_path)
            new_sheet = self.modify.create_sheet("Fails")
            new_sheet["A1"] = "ID"
            new_sheet["B1"] = "Reason"
        else:
            self.modify = load_workbook(self.final_path)
            new_sheet = self.modify.get_sheet_by_name("Fails")
            self.modify.active = new_sheet #// activates the fails sheet
            max_row = self.modify.active.max_row

            for x in range(1, max_row+1):
                cell = self.modify.active.cell(row=x, column=1) #// column or row must be at least 1
                if cell.value == self.id:
                    found = True
                    break
        if not found:
            new_sheet.append([self.id, self.reason])
        else:
            print("already there")

        
        self.modify.save(self.raport_path+f"\{self.raport_name}.xlsx")
        
        


